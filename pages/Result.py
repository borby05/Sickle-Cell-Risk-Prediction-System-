import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import inch
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font

st.set_page_config(
    page_title="Prediction Report",
    page_icon="📋",
    layout="wide"
)

st.title("📋 Prediction Report")

if "prediction" not in st.session_state:

    st.warning("No prediction found.")

    st.info("Please perform a prediction first.")

    st.stop()

patient = st.session_state["patient_data"]

prediction = st.session_state["prediction"]

prob = st.session_state["probabilities"]

st.subheader("Risk Assessment")

if prediction == "Low":
    st.success(f"Predicted Risk Level: {prediction}")

elif prediction == "Medium":
    st.warning(f"Predicted Risk Level: {prediction}")

else:
    st.error(f"Predicted Risk Level: {prediction}")


st.subheader("Prediction Confidence")

labels = ["Low", "Medium", "High"]

for label, p in prob.items():

    st.write(f"**{label}:** {p*100:.2f}%")

    st.progress(float(p))

st.subheader("Patient Information")

st.dataframe(
    patient,
    use_container_width=True
)

st.subheader("Clinical Interpretation")

if prediction == "Low":

    st.success("""
The patient's laboratory values and clinical information indicate a low risk category.
Continue routine monitoring and maintain regular clinical follow up.
""")

elif prediction == "Medium":

    st.warning("""
The patient falls within the medium risk category.
Closer monitoring and periodic laboratory assessment are recommended.
""")

else:

    st.error("""
The patient is classified as high risk.
Immediate clinical evaluation and continuous monitoring are strongly recommended.
""")


def generate_pdf(patient, prediction, prob):

    # Convert DataFrame to dictionary
    if hasattr(patient, "iloc"):
        patient = patient.iloc[0].to_dict()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=(8.27 * inch, 11.69 * inch),
        topMargin=25,
        bottomMargin=25,
        leftMargin=30,
        rightMargin=30,
    )

    styles = getSampleStyleSheet()

    story = []

    title = Paragraph("<b><font size=20>Sickle Cell Risk Assessment Report</font></b>", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    date = Paragraph(
        f"Generated on: {datetime.now().strftime('%d %B %Y %I:%M %p')}",
        styles["Normal"],
    )

    story.append(date)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Patient Information</b>", styles["Heading2"]))
    story.append(Spacer(1, 8))

    patient_table = [["Feature", "Value"]]

    for key, value in patient.items():
        patient_table.append([str(key), str(value)])

    table = Table(patient_table, colWidths=[220, 220])

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),

        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))

    story.append(table)

    story.append(Spacer(1, 25))

    story.append(Paragraph("<b>Prediction Result</b>", styles["Heading2"]))
    story.append(Paragraph(f"Risk Level: <b>{prediction}</b>", styles["BodyText"]))

    story.append(Spacer(1, 18))

    story.append(Paragraph("<b>Prediction Confidence</b>", styles["Heading2"]))

    confidence_table = [
        ["Risk Level", "Probability"],
        ["Low", f"{prob['Low']:.2%}"],
        ["Medium", f"{prob['Medium']:.2%}"],
        ["High", f"{prob['High']:.2%}"],
    ]

    table2 = Table(confidence_table, colWidths=[220, 220])

    table2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))

    story.append(table2)

    doc.build(story)

    buffer.seek(0)

    return buffer


def generate_csv(patient, prediction, prob):

    report = patient.copy()

    report["Predicted Risk"] = prediction

    report["Low Probability"] = prob["Low"]

    report["Medium Probability"] = prob["Medium"]

    report["High Probability"] = prob["High"]

    df = pd.DataFrame(report)

    return df.to_csv(index=False).encode("utf-8")


def generate_excel(patient, prediction, prob):

    if hasattr(patient, "iloc"):
        patient = patient.iloc[0].to_dict()

    wb = Workbook()

    ws = wb.active

    ws.title = "Risk Assessment"

    ws["A1"] = "Feature"
    ws["B1"] = "Value"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    row = 2

    for key, value in patient.items():
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = value
        row += 1

    row += 2

    ws.cell(row=row, column=1).value = "Predicted Risk"
    ws.cell(row=row, column=2).value = prediction

    row += 2

    ws.cell(row=row, column=1).value = "Low Probability"
    ws.cell(row=row, column=2).value = float(prob["Low"])

    row += 1

    ws.cell(row=row, column=1).value = "Medium Probability"
    ws.cell(row=row, column=2).value = float(prob["Medium"])

    row += 1

    ws.cell(row=row, column=1).value = "High Probability"
    ws.cell(row=row, column=2).value = float(prob["High"])

    excel = BytesIO()

    wb.save(excel)

    excel.seek(0)

    return excel
pdf_file = generate_pdf(patient, prediction, prob)

csv_file = generate_csv(
    patient,
    prediction,
    prob)

excel_file = generate_excel(
    patient,
    prediction,
    prob
)

col1, col2, col3 = st.columns(3)

with col1:

    st.download_button(
        "📄 Download PDF",
        data=pdf_file,
        file_name="Sickle_Cell_Report.pdf",
        mime="application/pdf",
        use_container_width=True
    )

with col2:

    st.download_button(
        "📊 Download CSV",
        data=csv_file,
        file_name="Sickle_Cell_Report.csv",
        mime="text/csv",
        use_container_width=True
    )

with col3:

    st.download_button(
        "📗 Download Excel",
        data=excel_file,
        file_name="Sickle_Cell_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )